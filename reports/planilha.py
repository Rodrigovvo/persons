from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from person.models import Pessoa


class Planilha:
    """
    Classe para gerar planilha com informações das pessoas.
    """

    @classmethod
    def get_planilha(self, queryset=None):
        """
        Metodo que cria a planilha com informações das pessoas.
        retorna um BytesIO (planilha) e o título da planilha.
        return: BytesIO, str
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Pessoas"

        headers = ["Nome", "Email", "Data de Nascimento", "Ativo", "Valor"]
        ws.append(headers)

        # Estilização dos headers
        for i, _ in enumerate(headers, start=1):
            cell = ws[f"{get_column_letter(i)}1"]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

        if queryset:
            pessoas = queryset
        else:
            pessoas = Pessoa.objects.filter().order_by("data_nascimento")

        for pessoa in pessoas:
            ws.append(
                [
                    pessoa.nome,
                    pessoa.email,
                    (
                        pessoa.data_nascimento.strftime("%d-%m-%Y")
                        if pessoa.data_nascimento
                        else None
                    ),
                    "ativo" if pessoa.ativo else "inativo",
                    pessoa.valor,
                ]
            )

        # Redimensionamento das colunas
        largura_fixa = 20
        for i in range(1, len(headers) + 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = largura_fixa

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file, ws.title
