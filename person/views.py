from datetime import datetime

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.utils import IntegrityError
from django.http import FileResponse, JsonResponse
from django.views import View
from django.views.generic import TemplateView
from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_BOOL, get_type
from openpyxl.utils.exceptions import InvalidFileException

from reports.planilha import Planilha
from utils.normalize import normalize_active

from .constants import EXCEL_COLUMNS
from .models import Pessoa


class PlanilhaTestView(TemplateView):
    template_name = "planilha.html"


class UploadExcelView(View):
    def post(self, request, *args, **kwargs):
        """
        Recebe um arquivo .xlsx e processa as informações para salvar no banco de dados.
        Não serão criadas pessoas cujo email e data de nascimento sejam inválidos ou já existam no banco de dados.
        Pessoa com idade inferior a 18 anos não será ativa.
        """
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"message":"Arquivo ausente."}, status=400)

        if not file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return JsonResponse({"message":"Formato de arquivo inválido."}, status=400)

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            processed_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_nascimento = row[EXCEL_COLUMNS["data de nascimento"]]
                email = row[EXCEL_COLUMNS["email"]]
                ativo = True

                try:
                    validate_email(email)
                except ValidationError:
                    print(
                        f"Forma inválida para o email '{row[EXCEL_COLUMNS['nome']]}': {email}"
                    )
                    continue

                if isinstance(data_nascimento, str):
                    try:
                        data_nascimento = datetime.strptime(
                            data_nascimento, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        print(
                            f"Forma inválida para a data de nascimento'{row[EXCEL_COLUMNS['nome']]}': {data_nascimento}"
                        )
                        continue

                if get_type(TYPE_BOOL, row[EXCEL_COLUMNS["ativo"]]):
                    ativo = normalize_active(row[EXCEL_COLUMNS["ativo"]])

                pessoa = Pessoa(
                    nome=row[EXCEL_COLUMNS["nome"]],
                    email=email,
                    data_nascimento=data_nascimento,
                    ativo=ativo,
                )

                if not pessoa.idade or pessoa.idade < 18:
                    pessoa.ativo = False

                try:
                    pessoa.save()
                except IntegrityError:
                    print(f"Registro duplicado: {pessoa.nome}")
                    continue

                if pessoa.ativo:
                    processed_data.append(
                        {
                            "nome": pessoa.nome,
                            "email": pessoa.email,
                            "data de nascimento": (
                                pessoa.data_nascimento.strftime("%Y-%m-%d")
                                if pessoa.data_nascimento
                                else None
                            ),
                            "ativo": pessoa.ativo,
                            "valor": pessoa.valor,
                        }
                    )
            if not processed_data:
                return JsonResponse({"message": "Sem dados para retorno."}, status=204)
            return JsonResponse(processed_data, safe=False, status=200)
        
        except Exception as e:
            return JsonResponse(
                {"message": f"Erro no processamento do arquivo: {str(e)}."},
                status=400
            )


class DownloadPlanilhaView(View):
    def get(self, request, *args, **kwargs):
        planilha, titulo = Planilha.get_planilha()
        response = FileResponse(planilha, as_attachment=True, filename=f"{titulo}.xlsx")
        return response
