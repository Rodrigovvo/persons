import os
from datetime import date

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

from person.models import Pessoa


class UploadExcelViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse("upload-planilha")

    def test_no_file_provided(self):
        response = self.client.post(self.url)
        self.assertEqual(response.status_code, 400)
        self.assertIn(b"Arquivo ausente.", response.content)

    def test_invalid_file_upload(self):
        invalid_file = SimpleUploadedFile(
            "test.txt", b"Invalid content", content_type="text/plain"
        )
        response = self.client.post(self.url, {"file": invalid_file})
        self.assertEqual(response.status_code, 400)

    def test_valid_file_upload(self):
        file_path = os.path.join(
            os.path.dirname(__file__), "..", "Tabela Desafio Python - Agosto2024.xlsx"
        )

        with open(file_path, "rb") as file:
            upload_file = SimpleUploadedFile(
                "Tabela Desafio Python - Agosto2024.xlsx",
                file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response = self.client.post(self.url, {"file": upload_file})

        self.assertEqual(response.status_code, 200)
        self.assertIn(b'"nome"', response.content)


class DownloadPlanilhaViewTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse("download-planilha")
        Pessoa.objects.create(
            nome="João Silva",
            email="joao@example.com",
            data_nascimento=date(2008, 1, 1),
            ativo=True,
            valor=100.00,
        )
        Pessoa.objects.create(
            nome="Maria Souza",
            email="maria@example.com",
            data_nascimento=date(1985, 5, 5),
            ativo=True,
            valor=150.00,
        )

    def test_download_planilha(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(hasattr(response, "streaming_content"))

        from io import BytesIO

        from openpyxl import load_workbook

        excel_file = BytesIO(b"".join(response.streaming_content))
        wb = load_workbook(excel_file)
        ws = wb.active

        headers = ["Nome", "Email", "Data de Nascimento", "Ativo", "Valor"]
        self.assertEqual([cell.value for cell in ws[1]], headers)

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 2)
        self.assertEqual(
            data_rows[0], ("João Silva", "joao@example.com", "1990-01-01", True, 150.00)
        )
        self.assertEqual(
            data_rows[1],
            ("Maria Souza", "maria@example.com", "1985-05-05", True, 150.00),
        )
